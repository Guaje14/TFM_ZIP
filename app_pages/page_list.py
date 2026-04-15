# ============================================
# Page List
# ============================================

# Importar librerías 
import streamlit as st
import pandas as pd
import os
from PIL import Image
from openpyxl import load_workbook
import base64
from fpdf import FPDF
import streamlit.components.v1 as components

# Importar rutas de configuración
from common.config import (
    ASSETSIMG, DATA_DIR, ASSETSFONTS
)

# Importar funciones para generar plots
from common.plots import (
    plot_nationality_choropleth,
    plot_competitions_pie,
    plot_age_distribution,
    plot_positions
)

# Importar funciones de datos y filtros
from controllers.db_controller import load_stats_players_fbref
from controllers.user_controller import load_users
from common.pdf_utils import get_watermark
from common.filters import apply_player_filters_lineup_list
    
# Función que da cómo resultado la página List
def page_list():
    
    # Añadir CSS de impresión
    st.markdown(
    """
    <style>

    /* estilos solo cuando se imprime */
    @media print {
        
        body {
            zoom: 0.7;
        }

        /* ocultar sidebar */
        section[data-testid="stSidebar"] {
            display: none;
        }

        /* ocultar menú superior de streamlit */
        header {
            display: none;
        }

        /* ocultar footer */
        footer {
            display: none;
        }

        /* expandir contenido principal */
        .main {
            margin-left: 0px;
            width: 100%;
        }

    }

    </style>
    """,
    unsafe_allow_html=True
    )
    
    # Crear columnas para centrar la imagen
    list_imgcol1, list_imgcol2, list_imgcol3 = st.columns([3, 2, 3])

    with list_imgcol2:

        # Cargar imagen desde assets
        list_img_header = Image.open(ASSETSIMG / "Freepick List.png")
        
        # Mostrar imagen en Streamlit
        st.image(list_img_header, width=120)

    # Mostrar título de la página
    st.header("Player List")
    
    # Cargar datos de jugadores desde la base de datos
    list_df_players = load_stats_players_fbref()
    
    # Cargar usuarios disponibles
    list_users = load_users()
    list_users_list = [u.username for u in list_users]  
    
    # Definir constantes
    LISTS = ["GK", "DF", "LT", "MF", "EX", "FW"]
    EXCEL_FILE = DATA_DIR / "list_players_register.xlsx"
    pos_order = ["GK", "DF", "MF", "FW"]
    
    # Definir valores por defecto de session_state
    list_defaults = {

        "list_league_filter": "All",
        "list_team_filter": "All",
        "list_pos_filter": "All",
        "list_player_to_add": "Select",

        "list_user_select": "Select",
        "list_list_choice": "All",
        "list_comment": "",
        "list_note": "Target",

        "list_view_list": "All",
        "list_view_users_list": "All",

        "list_do_reset": False
    }
    
    # Inicializar session_state si no existe
    for key, value in list_defaults.items():
        st.session_state.setdefault(key, value)
    
    # Aplicar reset si está activado
    if st.session_state["list_do_reset"]:

        # Reiniciar filtros de selección
        st.session_state["list_league_filter"] = "All"
        st.session_state["list_team_filter"] = "All"
        st.session_state["list_pos_filter"] = "All"
        st.session_state["list_player_to_add"] = "Select"

        st.session_state["list_user_select"] = "Select"
        st.session_state["list_list_choice"] = "All"
        st.session_state["list_comment"] = ""
        st.session_state["list_note"] = "Target"

        st.session_state["list_view_list"] = "All"
        st.session_state["list_view_users_list"] = "All"

        # Desactivar flag de reset
        st.session_state["list_do_reset"] = False

        # Recargar página
        st.rerun()

    # Función auxiliar para cargar hoja de Excel
    def load_sheet(sheet_name):

        # Definir columnas esperadas
        columns = ["Player", "User", "League", "Team", "Position", "List", "Comment", "Note"]

        # Devolver DataFrame vacío si no existe archivo
        if not os.path.exists(EXCEL_FILE):
            return pd.DataFrame(columns=columns)

        try:
            # Cargar hoja solicitada
            return pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        except:
            # Devolver DataFrame vacío si falla la lectura
            return pd.DataFrame(columns=columns)
        
    # Función auxiliar para guardar jugador en Excel
    def save_player_to_excel(data, sheet_name):

        # Convertir datos a DataFrame
        df_new = pd.DataFrame([data])

        # Crear archivo si no existe
        if not os.path.exists(EXCEL_FILE):

            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
                df_new.to_excel(writer, sheet_name=sheet_name, index=False)

            return

        # Cargar libro existente
        book = load_workbook(EXCEL_FILE)

        # Si la hoja existe, concatenar datos
        if sheet_name in book.sheetnames:

            df_old = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
            df_final = pd.concat([df_old, df_new], ignore_index=True)

        else:
            # Si no existe la hoja, crearla
            df_final = df_new

        # Guardar hoja actualizada
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_final.to_excel(writer, sheet_name=sheet_name, index=False)

    def delete_player(sheet_name, player_row):

        # No hacer nada si no existe archivo
        if not os.path.exists(EXCEL_FILE):
            return

        # Cargar datos de la hoja
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)

        # Rellenar NaN para comparar bien
        df = df.fillna("")
        player_row = {k: ("" if pd.isna(v) else v) for k, v in player_row.items()}

        # Crear máscara para identificar exactamente la fila a eliminar
        mask = (
            (df["Player"] == player_row["Player"]) &
            (df["User"] == player_row["User"]) &
            (df["League"] == player_row["League"]) &
            (df["Team"] == player_row["Team"]) &
            (df["Position"] == player_row["Position"]) &
            (df["List"] == player_row["List"]) &
            (df["Comment"] == player_row["Comment"]) &
            (df["Note"] == player_row["Note"])
        )

        # Eliminar solo la primera coincidencia exacta
        index_to_drop = df[mask].index

        if len(index_to_drop) > 0:
            df = df.drop(index=index_to_drop[0])

        # Guardar hoja actualizada
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
    # Crear layout principal
    form_col, col_viewer = st.columns([1, 2])
            
    # Construir formulario de jugador
    with form_col:
        
        st.subheader("Add Player")

        # Aplicar filtros reutilizables de liga, equipo, posición y jugador
        list_league, list_team, list_pos, list_player_to_add, list_players_filtered = apply_player_filters_lineup_list(
            df=list_df_players,
            league_key="list_league_filter",
            team_key="list_team_filter",
            pos_key="list_pos_filter",
            player_key="list_player_to_add",
            pos_order=pos_order
        )

        # Crear formulario de scouting
        with st.form("player_form"):
            
            st.write("Add Scout Data")
            
            # Seleccionar usuario
            user_selected = st.selectbox(
                "User",
                ["Select"] + list_users_list,
                key="list_user_select"
            )
            
            # Seleccionar lista
            list_choice = st.selectbox(
                "List",
                ["Select"] + LISTS,
                key="list_list_choice"
            )

            # Añadir comentario
            comment = st.text_area(
                "Comment (optional)",
                key="list_comment"
            )

            # Seleccionar rating
            note = st.radio(
                "Scout Rating",
                ["Target", "Watch", "Prospect"],
                key="list_note"
            )

            # Enviar formulario
            submitted = st.form_submit_button("Add Player")

            # Procesar envío
            if submitted:

                if list_league == "All":
                    st.warning("Please select a league.")
                elif list_team == "All":
                    st.warning("Please select a team.")
                elif list_pos == "All":
                    st.warning("Please select a position.")
                elif list_player_to_add == "Select":
                    st.warning("Please select a player.")
                elif user_selected == "Select":
                    st.warning("Please select a user.")
                elif list_choice == "Select":
                    st.warning("Please select a list.")
                else:
                    player_data = {
                        "Player": list_player_to_add,
                        "User": user_selected,
                        "League": list_league,
                        "Team": list_team,
                        "Position": list_pos,
                        "List": list_choice,
                        "Comment": comment,
                        "Note": note
                    }

                    save_player_to_excel(player_data, list_choice)
                    st.success(f"Player saved: {list_player_to_add}")
    
    # Crear botón de reset
    if st.button("🔄 Reset Filters"):

        # Activar flag de reset en session_state
        st.session_state["list_do_reset"] = True
        st.rerun()
        
    # Construir sección de visualización de listas
    with col_viewer:

        # Mostrar título de la sección
        st.subheader("Player Lists")

        # Crear columnas para filtros de visualización
        list_col1, list_col2 = st.columns(2)

        with list_col1:

            # Seleccionar lista de scouting
            selected_list = st.selectbox(
                "Select list",
                ["All"] + LISTS,
                key="list_view_list"
            )

        with list_col2:

            # Seleccionar usuario creador de la lista
            usar_list = st.selectbox(
                "User list",
                ["All"] + list_users_list,
                key="list_view_users_list"
            )

        # Cargar todas las listas si se selecciona "All"
        if selected_list == "All":

            dfs = []

            # Cargar cada lista individual
            for l in LISTS:

                df_temp = load_sheet(l)
                df_temp["List"] = l  # Añadir identificador de lista
                dfs.append(df_temp)

            # Unir todas las listas en un único DataFrame
            df_list = pd.concat(dfs, ignore_index=True)

        else:

            # Cargar lista específica seleccionada
            df_list = load_sheet(selected_list)

        # Filtrar por usuario si se selecciona uno específico
        if usar_list != "All":

            df_list = df_list[df_list["User"] == usar_list]

        # Mostrar mensaje si no hay datos
        if df_list.empty:

            st.info("No players in this list")

        else:
            
            # Definir límite de elementos por página
            display_limit = 8

            # Inicializar índice de paginación si no existe
            if "page_idx" not in st.session_state:
                st.session_state["page_idx"] = 0

            # Calcular número total de páginas
            total_pages = (len(df_list) - 1) // display_limit + 1

            # Crear botones de navegación
            nav_col1, nav_col2, nav_col3 = st.columns([2, 4, 1])

            with nav_col1:

                # Retroceder página si es posible
                if st.button("⬅️ Previous") and st.session_state["page_idx"] > 0:
                    st.session_state["page_idx"] -= 1

            with nav_col3:

                # Avanzar página si es posible
                if st.button("Next ➡️") and st.session_state["page_idx"] < total_pages - 1:
                    st.session_state["page_idx"] += 1

            # Calcular rango de elementos a mostrar
            start_idx = st.session_state["page_idx"] * display_limit
            end_idx = start_idx + display_limit
            df_display = df_list.iloc[start_idx:end_idx]

            # Mostrar mensaje si la página está vacía
            if df_display.empty:

                st.info("No players in this page")

            else:

                # Renderizar jugadores de la página actual
                for i, row in df_display.iterrows():

                    player_register_c1, player_register_c2 = st.columns([12, 1])

                    with player_register_c1:

                        # Mostrar información del jugador
                        st.markdown(
                            f"""
                            **{row['Player']}** | 🏟️ {row['Team']} | 🏆 {row['League']}  
                            📝 {row['Comment']}  
                            ⭐ **{row['Note']}** | 👤 **{row['User']}**
                            """
                        )

                    with player_register_c2:

                        # Eliminar jugador de la lista
                        if st.button("🗑️", key=f"del_{row['Player']}_{start_idx+i}"):

                            delete_player(row["List"], row.to_dict())
                            st.rerun()

            # Mostrar información de paginación
            st.info(f"Page {st.session_state['page_idx'] + 1} of {total_pages}")

    # Renderizar botón de impresión
    components.html(
    """
    <button onclick="parent.window.print()" style="
        padding:8px 14px;
        font-size:14px;
        cursor:pointer;
        background-color:#2563eb;
        color:white;
        border:none;
        border-radius:6px;">
        🖨️ Print Page
    </button>
    """,
    height=55
    )
    
    # Verificar que el DataFrame no esté vacío
    if not df_list.empty:

        # Crear botón para generar PDF
        if st.button("⚙️ Prepare PDF"):

            # Crear documento PDF
            pdf_list = FPDF()
            pdf_list.add_page()

            # Añadir fuente personalizada
            pdf_list.add_font("DejaVu", "", str(ASSETSFONTS / "DejaVuSans.ttf"), uni=True)

            # Escribir título del PDF
            pdf_list.set_font("DejaVu", "", 35)
            pdf_list.cell(0, 12, "List of Players", ln=True, align="C")
            pdf_list.ln(10)

            # Definir cabecera de tabla
            pdf_list.set_font("DejaVu", "", 12)
            col_widths_list = [35, 35, 35, 20, 20, 20, 20]
            headers_list = ["Player", "Team", "League", "Position", "List", "Note", "User"]

            # Crear cabecera de tabla
            for i, header in enumerate(headers_list):
                pdf_list.cell(col_widths_list[i], 8, header, border=1, align="C")
            pdf_list.ln()

            # Definir función para truncar texto largo
            def truncate(text, max_len=19):
                return text[:max_len] + "..." if len(text) > max_len else text

            # Definir fuente para datos
            pdf_list.set_font("DejaVu", "", 8)

            # Recorrer filas del DataFrame
            for _, row in df_list.iterrows():

                pdf_list.cell(col_widths_list[0], 8, truncate(str(row["Player"])), border=1)
                pdf_list.cell(col_widths_list[1], 8, truncate(str(row["Team"])), border=1)
                pdf_list.cell(col_widths_list[2], 8, truncate(str(row["League"])), border=1)
                pdf_list.cell(col_widths_list[3], 8, str(row["Position"]), border=1)
                pdf_list.cell(col_widths_list[4], 8, str(row["List"]), border=1)
                pdf_list.cell(col_widths_list[5], 8, truncate(str(row["Note"])), border=1)
                pdf_list.cell(col_widths_list[6], 8, str(row["User"]), border=1)
                pdf_list.ln()

            # Añadir marca de agua al PDF
            logo_buffer_list = get_watermark(alpha=10, user_obj=st.session_state.get("user"))
            pdf_list.image(logo_buffer_list, x=55, y=100, w=100)

            # Generar PDF en memoria
            pdf_bytes_list = pdf_list.output(dest="S")  

            # Codificar PDF en base64
            b64_list = base64.b64encode(pdf_bytes_list).decode()

            # Crear botón de descarga del PDF
            components.html(
                f"""
                <a href="data:application/pdf;base64,{b64_list}" download="player_list.pdf">
                    <button style="
                        padding:8px 14px;
                        font-size:14px;
                        cursor:pointer;
                        background-color:#dc2626;
                        color:white;
                        border:none;
                        border-radius:6px;">
                        📄 Export to PDF
                    </button>
                </a>
                """,
                height=55
            )

    # Separar sección visual
    st.divider()

    # Renombrar columnas para análisis
    df_list_renamed = df_list.rename(columns={
        "League": "stats_Comp",
        "Team": "stats_Squad",
        "Position": "stats_Pos"
    })

    # Unir dataset principal con lista de scouting
    df_list_merge = pd.merge(
        list_df_players,
        df_list_renamed,
        on=["Player", "stats_Comp", "stats_Squad", "stats_Pos"],
        how="inner"
    )

    # Eliminar duplicados por jugador y equipo
    df_list_merge_cleaned = df_list_merge.drop_duplicates(subset=["Player", "stats_Squad"])

    # Generar mapa de nacionalidades
    fig_map = plot_nationality_choropleth(df_list_merge_cleaned)
    st.plotly_chart(fig_map, use_container_width=True)

    # Separar visualización
    st.divider()

    # Crear layout de gráficos
    plots_ovierview_col1, plots_ovierview_col2, plots_ovierview_col3, plots_ovierview_col4, plots_ovierview_col5 = st.columns([2.3, 0.7, 2, 0.7, 2.5])

    # Mostrar distribución de edades
    with plots_ovierview_col1:
        fig_age = plot_age_distribution(df_list_merge_cleaned)
        st.plotly_chart(fig_age, use_container_width=True)

    # Mostrar distribución por posición
    with plots_ovierview_col3:
        fig_pos = plot_positions(df_list_merge_cleaned)
        st.plotly_chart(fig_pos, use_container_width=True)

    # Mostrar distribución por competición
    with plots_ovierview_col5:
        fig_comp = plot_competitions_pie(df_list_merge_cleaned)
        st.plotly_chart(fig_comp, use_container_width=True)